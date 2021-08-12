import openpyxl
import math
import decision_tree
from anytree import Node
from anytree import NodeMixin, RenderTree
from anytree.exporter import DotExporter
from graphviz import Source
from graphviz import render
from anytree.dotexport import RenderTreeGraph
import cv2
import os

os.environ["PATH"] += os.pathsep + 'C:/Program Files/Graphviz/bin'



def index_calculator(symptom , attribute):

    index_list = []
    start = 0
    for i in range(len(symptom)):
        try:
            index = symptom.index(attribute, start, len(symptom))
            index_list.append(index)
            start = index + 1
        except:
            break
    return index_list


def TreeTraverse(node, index_list, weight):

    number = dict.get(node.name)
    symp = symptomss_list[number]
    result = 'sss'
    if node.parent == None:
        for i in index_list:
            if symp[i] == weight:

                return diagnosis[i]


    else:
        print(node.name)
        print(node.parent.name)
        new_index = []
        for i in index_list:
            if symp[i] == weight:
                new_index.append(i)
        hold_weight = node.weight

        node = node.parent
        result = TreeTraverse(node,new_index, hold_weight)
        return result






def ID3_algorithm(current_node_index,symptoms_list,symptom_names,parent):



    entropy_list = []

    new_symptoms_list = []
    new_symptom_names = []
    for name in symptom_names:
        if name == symptom_names[current_node_index]:
            continue
        else:
            new_symptom_names.append(name)



    hold = []
    index_list = []
    unique_attributes = set(symptoms_list[current_node_index])
    for attribute in unique_attributes:
        new_symptoms_list.clear()

        index_list = index_calculator(symptoms_list[current_node_index], attribute)
        for symptom in symptoms_list:
            if symptom ==  symptoms_list[current_node_index] :
                continue

            hold.clear()
            for i in index_list:
                hold.append(symptom[i])
            temp = hold.copy()
            new_symptoms_list.append(temp)
        entropy_list = make_entropy_list(new_symptoms_list , index_list)
        lowest_entropy_index = entropy_list.index(min(entropy_list))



        total_num = list(range(0,10))

        if sum(entropy_list) == 0 :

            result = TreeTraverse(parent, total_num, attribute)
            leaf = decision_tree.WNode(result, parent=parent, weight=attribute)
        else:
            new_node = decision_tree.WNode(new_symptom_names[lowest_entropy_index] , parent=parent , weight=attribute)
            ID3_algorithm(lowest_entropy_index, new_symptoms_list,new_symptom_names ,new_node)




def make_entropy_list(symptoms_list, ind_list):
    entropy_list = []
    for symptom in symptoms_list:

        entropy = 0
        unique_values = set(symptom)
        for attribute in unique_values:
            index_list = []
            start = 0
            for i in range(len(symptom)):
                try:
                    index = symptom.index(attribute, start, len(symptom))
                    index_list.append(index)
                    start = index + 1
                except:
                    break

            entropy += calculateEntropy(symptom, attribute, ind_list )

        entropy_list.append(entropy)
    return entropy_list

def calculateEntropy (symp, attr, i_list):
    cov_num = 0
    cold_num = 0
    flu_num = 0

    Na = 0
    Ns = len(i_list)
    Cik_covid = 0
    Cik_cold = 0
    Cik_flu = 0
    unique_values = set(symp)
    for i in range(len(i_list)):
        if symp[i] == attr:
            Na += 1

    for i in range(len(i_list)):
        if symp[i] == attr:
            if diagnosis[i_list[i]] == 'Covid-19':
                Cik_covid += 1
            elif diagnosis[i_list[i]] == 'Cold':
                Cik_cold += 1
            else:
                Cik_flu += 1

    if len(symp) == 10:
        Pakj = len(i_list)/len(symp)
    else:
        Pakj= Na/Ns


    for i  in i_list:
        if diagnosis[i] == 'Covid-19':
            cov_num += 1
        elif diagnosis[i] == 'Cold':
            cold_num += 1
        elif diagnosis[i] == 'Flu':
            flu_num += 1



    if len(i_list)> 0 and len(symp) == 10:
        PCiAkj_covid = cov_num / len(i_list)
        PCiAkj_cold = cold_num / len(i_list)
        PCiAkj_flu = flu_num / len(i_list)
    elif len(i_list)> 0 and len(symp) < 10:
        PCiAkj_covid = Cik_covid / Na
        PCiAkj_cold = Cik_cold / Na
        PCiAkj_flu = Cik_flu / Na
    else:
        PCiAkj_covid = 0
        PCiAkj_cold = 0
        PCiAkj_flu = 0


    Pciakj_list = []
    Pciakj_list.append(PCiAkj_covid)
    Pciakj_list.append(PCiAkj_cold)
    Pciakj_list.append(PCiAkj_flu)

    entropy = 0
    for i in range(len(set(diagnosis))):

        if (Pciakj_list[i] > 0):
            entropy += -Pciakj_list[i] * math.log(Pciakj_list[i], 2)


    entropy *= Pakj

    return entropy



path = ('D:/University/3992/Artificial Intelligence/Proj/AI_Final/Book1.xlsx')

wb = openpyxl.load_workbook(path)
sheet = wb.active

fever = []
cough = []
muscle_pain = []
breath_shortness = []
sneeze_drops = []
gender = []
age = []
diagnosis = []

symptom_names = []
for column in sheet.iter_cols(min_col=2,max_col=8,min_row=1 , max_row=1):
    for cell in column:
        symptom_names.append(cell.value)



i=1
for column in sheet.iter_cols(min_col=2, min_row=2):
    for cell in column:
        if i == 1:
            fever.append(cell.value)
        elif i == 2:
            cough.append(cell.value)
        elif i == 3:
            muscle_pain.append(cell.value)
        elif i == 4:
            breath_shortness.append(cell.value)
        elif i == 5:
            sneeze_drops.append(cell.value)
        elif i == 6:
            gender.append(cell.value)
        elif i == 7:
            age.append(cell.value)
        elif i == 8:
            diagnosis.append(cell.value)
    i += 1
rows = []
horizontal_table = []

for i  in range(len(diagnosis)):
    rows.clear()
    rows.append(fever[i])
    rows.append(cough[i])
    rows.append(muscle_pain[i])
    rows.append(breath_shortness[i])
    rows.append(sneeze_drops[i])
    rows.append(gender[i])
    rows.append(age[i])
    rows.append(diagnosis[i])
    J = rows.copy()
    horizontal_table.append(J)


symptomss_list = []
symptomss_list.append(fever)
symptomss_list.append(cough)
symptomss_list.append(muscle_pain)
symptomss_list.append(breath_shortness)
symptomss_list.append(sneeze_drops)
symptomss_list.append(gender)
symptomss_list.append(age)


dict = {'Fever': 0,
        'Cough': 1,
        'Muscle Pain': 2,
        'Breath Shortness': 3,
        'Sneeze and drops': 4,
        'Gender': 5,
        'Age': 6}


### First: Make Entropy List For All the symptomss
entropy = 0
entropy_list = []
for symptom in symptomss_list:
    entropy = 0
    uniquee = set(symptom)
    for attr in uniquee:
        index_list = index_calculator(symptom , attr)
        entropy += calculateEntropy(symptom , attr, index_list)

    entropy_list.append(entropy)
### Picking the symptom with the least Entropy and removing it from the list of symptoms

lowest_entropy_index = entropy_list.index(min(entropy_list))
print(entropy_list)
#selected_node = symptoms_list[lowest_entropy_index]


### TODO: add the symptom  to the decision tree
root = decision_tree.WNode(symptom_names[lowest_entropy_index])


ID3_algorithm(lowest_entropy_index,symptomss_list, symptom_names, root)

for pre, _, node in RenderTree(root):
    print("%s%s (%s)" % (pre, node.name, node.weight or 0))




DotExporter(root,
            nodenamefunc=lambda node: f"{node.name}.{node.weight[0]}"if node.weight is not None else f"{node.name}" ,
            edgeattrfunc=lambda parent, child: "style=bold,label=%s" % (child.weight or 0)).to_picture("decision_tree.png")

image = cv2.imread('decision_tree.png',cv2.IMREAD_UNCHANGED)
cv2.imshow('Decision Tree',image)
cv2.waitKey(0)
